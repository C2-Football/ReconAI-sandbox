#!/usr/bin/env python3
"""Build the DHQ NFL-Fit / depth-fix / FantasyCalc comparison workbook."""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REP = os.path.join(ROOT, 'reports')
rows = json.load(open(os.path.join(REP, 'dhq-nfl-fit-rows.json')))
summ = json.load(open(os.path.join(REP, 'dhq-nfl-fit-diff-summary.json')))
OUT = os.path.join(REP, 'DHQ_NFL_Fit_BeforeAfter.xlsx')

HEADERS = ['pid', 'name', 'pos', 'team', 'age', 'fc_value', 'fc_rank', 'fc_scaled_today',
           'dhq_today', 'dhq_depthfix', 'dhq_depthfix_nflfit', 'dev_today_pct', 'dev_depthfix_pct',
           'dev_combined_pct', 'depthfix_delta', 'nflfit_delta', 'nfl_fit_label', 'role', 'opportunity', 'sitMult_after']
LABELS = {
    'fc_value': 'FC value', 'fc_rank': 'FC rank', 'fc_scaled_today': 'FC (scaled)',
    'dhq_today': 'DHQ today (live)', 'dhq_depthfix': 'DHQ +depth fix',
    'dhq_depthfix_nflfit': 'DHQ +depth +NFL-Fit',
    'dev_today_pct': 'today vs FC %', 'dev_depthfix_pct': '+depth vs FC %', 'dev_combined_pct': 'combined vs FC %',
    'depthfix_delta': 'Δ depth fix', 'nflfit_delta': 'Δ NFL-Fit', 'nfl_fit_label': 'NFL-Fit label',
    'sitMult_after': 'sitMult',
}
wb = openpyxl.Workbook()
hdr_fill = PatternFill('solid', fgColor='1F2937'); hdr_font = Font(color='FFFFFF', bold=True)
up = PatternFill('solid', fgColor='E8F5E9'); down = PatternFill('solid', fgColor='FDECEA')
under = PatternFill('solid', fgColor='FFF3CD')   # DHQ well below market


def scalar(v):
    return v if isinstance(v, (str, int, float)) or v is None else json.dumps(v, ensure_ascii=False)


def sheet(ws, data, title, color_key=None, mark_under_market=False):
    ws.title = title
    ws.append([LABELS.get(h, h) for h in HEADERS])
    for c in range(1, len(HEADERS) + 1):
        ws.cell(1, c).fill = hdr_fill; ws.cell(1, c).font = hdr_font
    for r in data:
        ws.append([r.get(h, '') for h in HEADERS])
        rr = ws.max_row
        if mark_under_market and isinstance(r.get('dev_combined_pct'), (int, float)) and r['dev_combined_pct'] <= -25:
            for c in range(1, len(HEADERS) + 1):
                ws.cell(rr, c).fill = under
        elif color_key:
            d = r.get(color_key, 0) or 0
            if d:
                fill = up if d > 0 else down
                for c in range(1, len(HEADERS) + 1):
                    ws.cell(rr, c).fill = fill
    ws.freeze_panes = 'C2'
    widths = {'name': 22, 'nfl_fit_label': 22, 'role': 8, 'opportunity': 24, 'team': 6, 'pid': 9}
    for i, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 12)


# Summary
s = wb.active; s.title = 'Summary'
s['A1'] = 'DHQ vs FantasyCalc — depth-chart fix + NFL-Fit layer'
s['A1'].font = Font(bold=True, size=14)
r = 3
for k, v in summ['summary'].items():
    if isinstance(v, dict):
        s.cell(r, 1, k).font = Font(bold=True); r += 1
        for kk, vv in v.items():
            s.cell(r, 1, '  ' + kk); s.cell(r, 2, scalar(vv)); r += 1
    else:
        s.cell(r, 1, k); s.cell(r, 2, scalar(v)); r += 1
s.column_dimensions['A'].width = 30; s.column_dimensions['B'].width = 95

# FC comparison — sorted by FC value (great players first), highlight DHQ ≥25% under market
fc_sorted = sorted([x for x in rows if x.get('fc_value', 0) > 0], key=lambda x: -x['fc_value'])
sheet(wb.create_sheet(), fc_sorted, 'vs FantasyCalc', mark_under_market=True)
# All players (by DHQ), NFL-Fit changes, depth-fix impact
sheet(wb.create_sheet(), rows, 'All players', 'nflfit_delta')
sheet(wb.create_sheet(), [x for x in rows if x.get('nflfit_delta')], 'NFL-Fit changed', 'nflfit_delta')
sheet(wb.create_sheet(), sorted(rows, key=lambda x: -abs(x.get('depthfix_delta', 0))), 'Depth-fix impact', 'depthfix_delta')
wb.save(OUT)
print('Saved', OUT, '—', len(rows), 'players;', len(fc_sorted), 'with FC values')
